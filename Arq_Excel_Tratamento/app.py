from openpyxl import load_workbook, Workbook
import re

# Arquivo de entrada
arquivo_entrada = "CodCli e NomCli.xlsx"

# Arquivo que será criado
arquivo_saida = "CodCli_NomCli_Separado.xlsx"

# ---------------------------------------------------------
# Abre o Excel original
# ---------------------------------------------------------
wb_entrada = load_workbook(arquivo_entrada)

# Pega explicitamente a primeira planilha
ws_entrada = wb_entrada.worksheets[0]

# ---------------------------------------------------------
# Cria o novo Excel
# ---------------------------------------------------------
wb_saida = Workbook()

# Pega explicitamente a primeira planilha criada
ws_saida = wb_saida.worksheets[0]

ws_saida.title = "Clientes"

# Cabeçalhos
ws_saida["A1"] = "CodCli"
ws_saida["B1"] = "NomCli"

linha_saida = 2

# ---------------------------------------------------------
# Percorre a primeira coluna do Excel original
# ---------------------------------------------------------
for linha in ws_entrada.iter_rows(
    min_col=1,
    max_col=1,
    values_only=True
):

    conteudo = linha[0]

    # Ignora linhas vazias
    if conteudo is None:
        continue

    conteudo = str(conteudo).strip()

    # Procura:
    # número no início + espaços + nome do cliente
    resultado = re.match(r"^(\d+)\s+(.+)$", conteudo)

    if resultado:

        codcli = int(resultado.group(1))
        nomcli = resultado.group(2).strip()

        ws_saida.cell(
            row=linha_saida,
            column=1,
            value=codcli
        )

        ws_saida.cell(
            row=linha_saida,
            column=2,
            value=nomcli
        )

        linha_saida += 1

# ---------------------------------------------------------
# Ajusta largura das colunas
# ---------------------------------------------------------
ws_saida.column_dimensions["A"].width = 12
ws_saida.column_dimensions["B"].width = 80

# ---------------------------------------------------------
# Salva o novo Excel
# ---------------------------------------------------------
wb_saida.save(arquivo_saida)

print("Arquivo criado com sucesso!")
print(f"Total de clientes: {linha_saida - 2}")
print(f"Arquivo: {arquivo_saida}")